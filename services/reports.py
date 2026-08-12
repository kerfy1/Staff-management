"""Генерація файлів: звіт по персоналу та бланк для перевірки зарплати.

Принципи:
- у бланку зарплати всі підсумки — ФОРМУЛИ, не захардкоджені числа,
  щоб керівник міг змінити ставку/години і одразу побачити перерахунок;
- жовта заливка = комірки, які керівник може правити руками;
- шрифт Arial, легенда на першому аркуші.
"""
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import config
import db
from services import dates, payroll

HEAD_FILL = PatternFill("solid", fgColor="1F4E79")
HEAD_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BASE_FONT = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")   # комірки для ручного редагування
INPUT_FONT = Font(name="Arial", size=10, color="0000FF")
MONEY = '#,##0.00'
BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)

KIND_UA = {"regular": "Звичайна", "extra": "Додаткова"}
REQ_UA = {"dayoff": "Вихідний", "sick": "Лікарняний", "extra_shift": "Додаткова зміна"}
STATUS_UA = {"pending": "На розгляді", "approved": "Схвалено", "rejected": "Відхилено"}


def _ensure_dir() -> str:
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)
    return config.OUTPUT_DIR


def _header(ws, headers: list[str], widths: list[int], row: int = 1) -> None:
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill, c.font = HEAD_FILL, HEAD_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _write(ws, row: int, values: list, money_cols: tuple[int, ...] = ()) -> None:
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = BASE_FONT
        c.border = BORDER
        if i in money_cols:
            c.number_format = MONEY


async def build_report(date_from: str, date_to: str, group_id: int | None = None) -> str:
    """Звіт: зведення по людях, всі зміни, заявки, відсутності."""
    rows = await payroll.calculate(date_from, date_to, group_id=group_id)
    shifts = await db.shifts_between(date_from, date_to, group_id=group_id)
    requests = await db.list_requests(status=None, date_from=date_from, date_to=date_to)
    absences = await db.absences_between(date_from, date_to)

    wb = Workbook()

    # --- Зведення ---
    ws = wb.active
    ws.title = "Зведення"
    ws["A1"] = f"Звіт по персоналу за період {dates.period_title(date_from, date_to)}"
    ws["A1"].font = Font(name="Arial", bold=True, size=13)
    ws["A2"] = f"Сформовано: {dates.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font = Font(name="Arial", size=9, italic=True)

    headers = ["ПІБ", "Група", "Днів", "Год. звич.", "Год. дод.", "Всього год.",
               "Ставка", "Лікарняні (дн)", "Вихідні (дн)", "Пропуски", f"Нараховано, {config.CURRENCY}"]
    _header(ws, headers, [26, 16, 8, 11, 11, 12, 10, 14, 13, 10, 16], row=4)

    r = 5
    for p in rows:
        _write(ws, r, [p.full_name, p.group_name, p.days_worked, p.regular_hours, p.extra_hours,
                       p.total_hours, p.rate, p.sick_days, p.dayoff_days, p.absences, p.total],
               money_cols=(7, 11))
        r += 1
    if rows:
        ws.cell(row=r, column=1, value="РАЗОМ").font = BOLD
        for col in (3, 4, 5, 6, 11):
            L = get_column_letter(col)
            c = ws.cell(row=r, column=col, value=f"=SUM({L}5:{L}{r-1})")
            c.font, c.number_format = BOLD, (MONEY if col == 11 else '#,##0.0')

    # --- Зміни ---
    ws2 = wb.create_sheet("Зміни")
    _header(ws2, ["Дата", "ПІБ", "Група", "Початок", "Кінець", "Годин", "Тип", "Коментар"],
            [12, 26, 16, 10, 10, 9, 14, 30])
    r = 2
    for s in shifts:
        _write(ws2, r, [s["day"], s["full_name"], s.get("group_name") or "—", s.get("start_at") or "",
                        s.get("end_at") or "", s.get("hours") or 0,
                        KIND_UA.get(s["kind"], s["kind"]), s.get("comment") or ""])
        r += 1

    # --- Заявки ---
    ws3 = wb.create_sheet("Заявки")
    _header(ws3, ["ПІБ", "Тип", "З", "По", "Статус", "Коментар", "Створено"],
            [26, 18, 12, 12, 14, 30, 18])
    r = 2
    for q in requests:
        _write(ws3, r, [q["full_name"], REQ_UA.get(q["kind"], q["kind"]), q["date_from"],
                        q["date_to"], STATUS_UA.get(q["status"], q["status"]),
                        q.get("comment") or "", q["created_at"]])
        r += 1

    # --- Відсутності ---
    ws4 = wb.create_sheet("Відсутності")
    _header(ws4, ["Дата", "ПІБ", "Причина"], [12, 26, 40])
    r = 2
    for a in absences:
        _write(ws4, r, [a["day"], a["full_name"], a.get("reason") or ""])
        r += 1

    _ensure_dir()
    path = os.path.join(config.OUTPUT_DIR, f"zvit_{date_from}_{date_to}.xlsx")
    wb.save(path)
    return path


async def build_payroll_sheet(date_from: str, date_to: str, group_id: int | None = None) -> str:
    """Бланк під перевірку зарплати: всі суми — формули, вхідні комірки підсвічені."""
    rows = await payroll.calculate(date_from, date_to, group_id=group_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Бланк ЗП"

    ws["A1"] = f"Відомість нарахування зарплати: {dates.period_title(date_from, date_to)}"
    ws["A1"].font = Font(name="Arial", bold=True, size=13)
    ws["A2"] = ("Легенда: жовті комірки (сині цифри) — вхідні дані, їх можна правити вручну; "
                "чорні колонки рахуються формулами і перераховуються автоматично.")
    ws["A2"].font = Font(name="Arial", size=9, italic=True)

    # Параметри — окремі комірки, на які посилаються формули
    ws["A4"] = "Параметри розрахунку"
    ws["A4"].font = BOLD
    params = [
        ("Коеф. додаткової зміни", config.EXTRA_SHIFT_MULTIPLIER),
        ("Оплата лікарняного (частка)", config.SICK_PAY_RATE),
        ("Норма годин у зміні", config.SHIFT_NORM_HOURS),
    ]
    for i, (label, value) in enumerate(params, start=5):
        ws.cell(row=i, column=1, value=label).font = BASE_FONT
        c = ws.cell(row=i, column=2, value=value)
        c.font, c.fill, c.border = INPUT_FONT, INPUT_FILL, BORDER
    # B5 = коеф., B6 = лікарняний, B7 = норма годин

    head_row = 9
    headers = ["ПІБ", "Група", "Ставка/год", "Год. звич.", "Год. дод.", "Лікарняні, дн",
               "Оплата звич.", "Оплата дод.", "Оплата лікарняних", "Нараховано",
               "Коригування", "До виплати", "Перевірено (підпис)"]
    _header(ws, headers, [24, 14, 11, 11, 11, 13, 13, 13, 16, 13, 13, 13, 20], row=head_row)

    r = head_row + 1
    for p in rows:
        ws.cell(row=r, column=1, value=p.full_name).font = BASE_FONT
        ws.cell(row=r, column=2, value=p.group_name).font = BASE_FONT
        for col, val in ((3, p.rate), (4, p.regular_hours), (5, p.extra_hours), (6, p.sick_days)):
            c = ws.cell(row=r, column=col, value=val)
            c.font, c.fill, c.border = INPUT_FONT, INPUT_FILL, BORDER
            c.number_format = MONEY if col == 3 else '#,##0.0'
        formulas = {
            7: f"=C{r}*D{r}",                       # звичайні години
            8: f"=C{r}*E{r}*$B$5",                  # додаткові × коеф.
            9: f"=F{r}*$B$7*C{r}*$B$6",             # лікарняні
            10: f"=G{r}+H{r}+I{r}",                 # нараховано
            11: 0,                                  # коригування (премія/штраф) — руками
            12: f"=J{r}+K{r}",                      # до виплати
        }
        for col, f in formulas.items():
            c = ws.cell(row=r, column=col, value=f)
            c.number_format = MONEY
            c.border = BORDER
            if col == 11:
                c.font, c.fill = INPUT_FONT, INPUT_FILL
            else:
                c.font = BASE_FONT
        sign = ws.cell(row=r, column=13, value="")
        sign.border = BORDER
        r += 1

    if rows:
        ws.cell(row=r, column=1, value="РАЗОМ").font = BOLD
        for col in (7, 8, 9, 10, 11, 12):
            L = get_column_letter(col)
            c = ws.cell(row=r, column=col, value=f"=SUM({L}{head_row+1}:{L}{r-1})")
            c.font, c.number_format, c.border = BOLD, MONEY, BORDER

    ws.cell(row=r + 2, column=1,
            value="Джерело даних: зміни та схвалені заявки в базі бота за вказаний період.").font = \
        Font(name="Arial", size=9, italic=True)

    _ensure_dir()
    path = os.path.join(config.OUTPUT_DIR, f"blank_zp_{date_from}_{date_to}.xlsx")
    wb.save(path)
    return path
